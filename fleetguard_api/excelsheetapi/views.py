from django.conf import settings
from django.shortcuts import render
import os
import cv2
import re
import pandas as pd
from django.http import JsonResponse
from django.utils.decorators import method_decorator
from django.views import View
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from paddleocr import PaddleOCR
import pytesseract
from urllib.parse import unquote,urlparse
from urllib.parse import quote
import os
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

import paddle
# Initialize PaddleOCR
ocr = PaddleOCR(use_angle_cls=True)


def extract_main_dimension(text):
    # Extract the main dimension part (numerical part with optional ± tolerance)
    match = re.search(r'\d+(\.\d+)?\s*±?\s*\d*(\.\d+)?', text)
    if match:
        return match.group().replace(' ', '')
    return text.strip()


def extract_base_value(text):
    try:
        # Extract the numeric part along with any decimal points and ignore other characters
        match = re.search(r'(\d+(\.\d+)?)', text)
        if match:
            return match.group(1)
        return text  # Return full text if no numeric value found
    except Exception as e:
        print(f"Error extracting base value from '{text}': {e}")
        return text


def FINALcompare_dimensions(ffpl_text, cust_dict):
    matched_pairs = []
    unmatched_ffpl = []
    unmatched_cust = cust_dict.copy()

    for f_text in ffpl_text:
        f_main_dim = extract_main_dimension(f_text)
        f_base_value = extract_base_value(f_text)
        found_match = False
        for c_no, c_text in list(unmatched_cust.items()):  # Use list to avoid runtime changes
            c_main_dim = extract_main_dimension(c_text)
            c_base_value = extract_base_value(c_text)
            if f_main_dim == c_main_dim or f_base_value == c_base_value:
                matched_pairs.append((f_text, c_text, c_no))
                del unmatched_cust[c_no]
                found_match = True
                break

        if not found_match:
            matched_pairs.append((f_text, '', ''))
            unmatched_ffpl.append(f_text)

    for c_no, c_text in unmatched_cust.items():
        matched_pairs.append(('', c_text, c_no))

    return matched_pairs


def extract_path_from_url(url):
    parsed_url = urlparse(url)
    path = unquote(parsed_url.path)
    path = path.replace('\\', '/').lstrip('/')  # Replace backslashes and remove leading slashes
    return os.path.normpath(path)

# Ensure paths are relative to MEDIA_ROOT
def construct_full_path(url_path):
    if not url_path.startswith(settings.MEDIA_URL):
        return None
    media_root_relative_path = url_path[len(settings.MEDIA_URL):]
    return os.path.join(settings.MEDIA_ROOT, media_root_relative_path)

@method_decorator(csrf_exempt, name='dispatch')
class GenerateExcelView(View):
    def extract_text(self, image_path, yolo_output):
        image = cv2.imread(image_path)
        text_list = []
        notes_list = []
        with open(yolo_output, 'r') as f:
            lines = f.readlines()

        boxes = []
        for line in lines:
            parts = line.split()
            class_index = int(parts[0])
            x, y, w, h = map(float, parts[1:])
            top_left = (int((x - w/2) * image.shape[1]), int((y - h/2) * image.shape[0]))
            bottom_right = (int((x + w/2) * image.shape[1]), int((y + h/2) * image.shape[0]))
            boxes.append((class_index, top_left, bottom_right))

        boxes.sort(key=lambda box: (box[0], box[1][0]))

        for i, (class_index, top_left, bottom_right) in enumerate(boxes):
            cropped_img = image[int(top_left[1]):int(bottom_right[1]), int(top_left[0]):int(bottom_right[0])]
            # cropped_img = preprocess_image(cropped_img)

            if class_index != 1:
                result = ocr.ocr(cropped_img)
                text = ' '.join([word[1][0] for word in result[0]]) if result and result[0] else ""
                text_list.append(text)
            else:
                try:
                    if not cropped_img.size:
                        raise ValueError("Empty Notes")
                    text = pytesseract.image_to_string(cropped_img, lang='eng', config='--psm 6')
                    notes_list.append(text)
                except Exception as e:
                    print(f"Error processing image: {e}")
                    notes_list.append("ERROR")
        return text_list, notes_list

    def FINALsave_to_excel(self,ffpl_text, ffpl_notes, cust_text, cust_notes, img_num):
        cust_dict = {i + 1: text for i, text in enumerate(cust_text)}  # Pair CUST_NO with CUST_TEXT
        matched_pairs = FINALcompare_dimensions(ffpl_text, cust_dict)
        max_length = max(len(matched_pairs), len(ffpl_notes), len(cust_notes))
        max_ffpl_notes = len(ffpl_notes)
        max_cust_notes = len(cust_notes)
        red_font = Font(color="FF0000")
        # Prepare lists to match the DataFrame columns
        ffpl_text_ordered = [pair[0] for pair in matched_pairs] + [''] * (max_length - len(matched_pairs))
        cust_text_ordered = [pair[1] for pair in matched_pairs] + [''] * (max_length - len(matched_pairs))
        cust_no_ordered = [pair[2] for pair in matched_pairs] + [''] * (max_length - len(matched_pairs))
        ffpl_notes += [''] * (max_length - len(ffpl_notes))
        cust_notes += [''] * (max_length - len(cust_notes))

        # Create a DataFrame
        data = {'FFPL_NO': range(1, max_length + 1),
                'FFPL_TEXT': ffpl_text_ordered,
                'CUST_NO': cust_no_ordered,
                'CUST_TEXT': cust_text_ordered,
                'FFPL_NOTES': ffpl_notes,
                'CUST_NOTES': cust_notes}
        df = pd.DataFrame(data)

        try:
            # Define directory and filename
            excelsheet_dir = os.path.join(settings.MEDIA_ROOT, 'Excelsheet')
            # Create the directory if it doesn't exist
            os.makedirs(excelsheet_dir, exist_ok=True)
            filename = f'detected_text_{img_num}.xlsx'
            file_path = os.path.join(excelsheet_dir, filename)

            # Write the DataFrame to an Excel file
            df.to_excel(file_path, index=False)
            print(f"Excel file saved: {file_path}")

            # Open the Excel file and apply borders and formatting
            wb = load_workbook(file_path)
            sheet = wb.active

            # Set the desired column widths (adjust as necessary)
            column_widths = [5, 40, 5, 40, 60, 60]
            for i, width in enumerate(column_widths, start=1):
                column_letter = get_column_letter(i)
                sheet.column_dimensions[column_letter].width = width

            # Apply borders to all cells
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

            # Apply word wrap for FFPL_TEXT and CUST_TEXT columns
            ffpl_text_column = get_column_letter(2)  # FFPL_TEXT column
            cust_text_column = get_column_letter(4)  # CUST_TEXT column
            ffpl_notes_column = get_column_letter(5)  # FFPL_NOTES column
            cust_notes_column = get_column_letter(6)  # CUST_NOTES column

            for column in [ffpl_text_column, cust_text_column]:
                for cell in sheet[column]:
                    cell.alignment = Alignment(wrap_text=True)

            # Highlight matching rows with yellow color
            yellow_fill = PatternFill(start_color="FFDD00", end_color="FFDD00", fill_type="solid")
            light_blue_fill = PatternFill(start_color="ADD8E6", end_color="AADDE6", fill_type="solid")

            # for i, pair in enumerate(matched_pairs, start=2):  # Start from 2 to skip header row
            #     if pair[0] and pair[1]:  # Both FFPL_TEXT and CUST_TEXT are present
            #         for cell in [sheet[f'{ffpl_text_column}{i}'], sheet[f'{cust_text_column}{i}']]:
            #             cell.fill = yellow_fill
            #

            black_font = Font(color="000000")  # Black font

            for i, pair in enumerate(matched_pairs, start=2):  # Start from 2 to skip header row
                if pair[0] and pair[1]:  # Both FFPL_TEXT and CUST_TEXT are present
                    base_f_value = extract_base_value(pair[0])
                    base_c_value = extract_base_value(pair[1])
                    if base_f_value == base_c_value and pair[0] == pair[1]:
                        for cell in [sheet[f'{ffpl_text_column}{i}'], sheet[f'{cust_text_column}{i}']]:
                            cell.fill = light_blue_fill
                            cell.font = black_font
                    elif base_f_value == base_c_value:
                        for cell in [sheet[f'{ffpl_text_column}{i}'], sheet[f'{cust_text_column}{i}']]:
                            cell.fill = light_blue_fill
                            cell.font = red_font
                    else:
                        for cell in [sheet[f'{ffpl_text_column}{i}'], sheet[f'{cust_text_column}{i}']]:
                            cell.font = black_font
                    # Compare FFPL_NOTES and CUST_NOTES

            # Merge the last row for FFPL_NOTES and CUST_NOTES columns and align text to top
            max_row = len(df) + 1
            ffpl_notes_column = get_column_letter(5)  # FFPL_NOTES column
            cust_notes_column = get_column_letter(6)  # CUST_NOTES column
            ffpl_notes_range = f'{ffpl_notes_column}{max_ffpl_notes + 1}:{ffpl_notes_column}{max_row}'
            cust_notes_range = f'{cust_notes_column}{max_cust_notes + 1}:{cust_notes_column}{max_row}'
            sheet.merge_cells(ffpl_notes_range)
            sheet.merge_cells(cust_notes_range)
            for column in [ffpl_notes_column, cust_notes_column]:
                for cell in sheet[column]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            wb.save(file_path)

        except Exception as e:
            print(f"Error occurred while saving Excel file: {e}")
        # filename = f'detected_text_{img_num}.xlsx'
        # Save the DataFrame to an Excel file, apply formatting, etc.
        # Return the filename for response
        return filename


    @csrf_exempt
    def post(self, request):
        # Get the uploaded images
        ffpl_image = request.POST.get('ffpl_image')
        cust_image = request.POST.get('cust_image')
        ffpl_yolo_output = request.POST.get('ffpl_yolo_output')
        cust_yolo_output = request.POST.get('cust_yolo_output')
        if not ffpl_image or not cust_image or not ffpl_yolo_output or not cust_yolo_output:
            return JsonResponse({'error': 'All image and YOLO output paths are required'}, status=400)

        decoded_ffpl_path=unquote(ffpl_image)
        decoded_cust_path=unquote(cust_image)
        decoded_ffpl_yolo=unquote(ffpl_yolo_output)
        decoded_cust_yolo=unquote(cust_yolo_output)

        # Ensure the path starts with the MEDIA_ROOT to prevent directory traversal attacks
        if not decoded_ffpl_path.startswith(settings.MEDIA_URL):
            return JsonResponse({'error': 'Invalid FFPL image URL'}, status=400)
        media_root_relative_path = decoded_ffpl_path[len(settings.MEDIA_URL):]
        full_ffpl_path = os.path.join(settings.MEDIA_ROOT, media_root_relative_path)

        if not decoded_cust_path.startswith(settings.MEDIA_URL):
            return JsonResponse({'error': 'Invalid CUST image URL'}, status=400)
        media_root_relative_path = decoded_cust_path[len(settings.MEDIA_URL):]
        full_cust_path = os.path.join(settings.MEDIA_ROOT, media_root_relative_path)

        if not decoded_ffpl_yolo.startswith(settings.MEDIA_URL):
            return JsonResponse({'error': 'Invalid FFPL YOLO URL'}, status=400)
        media_root_relative_path = decoded_ffpl_yolo[len(settings.MEDIA_URL):]
        ffpl_yolo_output_path = os.path.join(settings.MEDIA_ROOT, media_root_relative_path)

        if not decoded_cust_yolo.startswith(settings.MEDIA_URL):
            return JsonResponse({'error': 'Invalid CUST YOLO URL'}, status=400)
        media_root_relative_path = decoded_cust_yolo[len(settings.MEDIA_URL):]
        cust_yolo_output_path = os.path.join(settings.MEDIA_ROOT, media_root_relative_path)

        # Ensure the 'excelsheet' directory exists
        excelsheet_dir = os.path.join(settings.MEDIA_ROOT, 'excelsheet')
        os.makedirs(excelsheet_dir, exist_ok=True)

        # Perform OCR on both images
        ffpl_text, ffpl_notes = self.extract_text(full_ffpl_path, ffpl_yolo_output_path)
        cust_text, cust_notes = self.extract_text(full_cust_path, cust_yolo_output_path)

        # Generate Excel
        img_num = len(os.listdir(settings.MEDIA_ROOT)) # Use current count of files to create unique filename
        excel_filename = self.FINALsave_to_excel(ffpl_text, ffpl_notes, cust_text, cust_notes, img_num)
        excel_file_path = os.path.join('excelsheet', excel_filename)  # Path relative to MEDIA_URL

        return JsonResponse({'excel_file': excel_file_path})